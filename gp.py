#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
import pandas as pd
import numpy as np
import time
import warnings
from datetime import datetime
from pathlib import Path

from sklearn.preprocessing import StandardScaler, LabelEncoder
import xgboost as xgb

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ==================== 配置 ====================
OUTPUT_DIR = Path("./jiuyan_analysis")
OUTPUT_DIR.mkdir(exist_ok=True)

# ==================== 数据获取 ====================
def get_zt_data():
    print("📡 步骤1：获取今日涨停数据...")
    today_date = datetime.now().strftime('%Y%m%d')
    url = "https://push2ex.eastmoney.com/getTopicZTPool"
    params = {
        'ut': '7eea3edcaed734bea9cbfc24409ed989',
        'dpt': 'wz.ztzt',
        'Pageindex': '0',
        'pagesize': '5000',
        'sort': 'fbt:asc',
        'date': today_date,
        '_': str(int(time.time() * 1000)),
    }
    headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"}
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=10)
        data = resp.json()
        return pd.DataFrame(data.get('data', {}).get('pool', [])), today_date
    except:
        return pd.DataFrame(), today_date

df_raw, today_date = get_zt_data()
if df_raw.empty:
    print("❌ 今日无数据"); exit()

# --- 数据清洗 ---
col_map = {
    'c': '股票代码', 'n': '股票名称', 'p': '最新价', 'zdp': '涨跌幅', 
    'fbt': '首次封板时间', 'lbc': '连板天数', 'zbc': '炸板次数', 
    'fund': '封单金额', 'hs': '换手率', 'hybk': '所属行业'
}
df = df_raw.rename(columns=col_map)
df['股票代码'] = df['股票代码'].astype(str).str.zfill(6)
df['最新价'] = pd.to_numeric(df['最新价'], errors='coerce') / 1000
df['涨跌幅'] = pd.to_numeric(df['涨跌幅'], errors='coerce') / 100
df['连板天数'] = pd.to_numeric(df['连板天数'], errors='coerce').fillna(1).astype(int)
df['封单金额'] = pd.to_numeric(df['封单金额'], errors='coerce').fillna(0) / 1e8
df['换手率'] = pd.to_numeric(df['换手率'], errors='coerce').fillna(0)
df['炸板次数'] = pd.to_numeric(df['炸板次数'], errors='coerce').fillna(0).astype(int)
df['所属行业'] = df['所属行业'].fillna('未知')

# 时间格式化修复：显示为 24小时格式 HH:mm
def fmt_time_fixed(raw):
    try:
        s = str(int(float(raw))).zfill(6)
        hh, mm = s[:2], s[2:4]
        # 简单校验小时合法性
        if int(hh) > 23: hh = "09"
        return f"{hh}:{mm}"
    except: return "09:30"
df['首次封板时间'] = df['首次封板时间'].apply(fmt_time_fixed)
df['连板状态'] = df['连板天数'].apply(lambda x: "首板" if x <= 1 else f"{x}天{x}板")

# ==================== 步骤2：模型评分 ====================
print("🤖 步骤2：模型评分...")
df['行业编码'] = LabelEncoder().fit_transform(df['所属行业'])
feat_cols = ['最新价', '涨跌幅', '连板天数', '行业编码', '换手率', '炸板次数', '封单金额']
X = df[feat_cols].fillna(0)
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)

model = xgb.XGBRegressor(n_estimators=100, max_depth=4, learning_rate=0.1, random_state=42)
model.fit(X_scaled, df['连板天数'])
df['预测强度'] = model.predict(X_scaled).round(2)
df['强度等级'] = pd.cut(df['预测强度'], bins=[-np.inf, 2, 4, 6, np.inf], labels=['弱','中','强','极强'])

# ==================== 步骤3：Excel 生成与格式化 ====================
wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def write_and_style(ws_name, target_df, custom_formats=None):
    ws = wb.create_sheet(ws_name)
    # 写入表头
    for c_idx, col in enumerate(target_df.columns, 1):
        cell = ws.cell(1, c_idx, col)
        cell.fill, cell.font, cell.border = header_fill, header_font, thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据
    for r_idx, row in enumerate(target_df.values, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            col_name = target_df.columns[c_idx-1]
            if custom_formats and col_name in custom_formats:
                cell.number_format = custom_formats[col_name]

    # 自适应列宽逻辑 (增强型)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            if cell.value is not None:
                # 处理百分比和特殊格式后的视觉长度
                if cell.number_format == '0.00%':
                    val_str = f"{cell.value * 100:.2f}%"
                elif "亿" in cell.number_format:
                    val_str = f"{cell.value:.2f}亿"
                else:
                    val_str = str(cell.value)
                
                # 计算显示宽度：中文算2.2 (稍微留宽一点), 英文算1.1
                display_width = sum(2.2 if '\u4e00' <= char <= '\u9fff' else 1.1 for char in val_str)
                if display_width > max_length:
                    max_length = display_width
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 65)

# --- 生成 Sheet ---

# 1. 今日涨停个股 (按连板天数降序, 格式化跌涨幅、封单、时间)
df_main = df[['股票代码','股票名称','最新价','涨跌幅','连板状态','连板天数','首次封板时间','所属行业','封单金额','预测强度','强度等级']]
df_main = df_main.sort_values('连板天数', ascending=False)
fmt_main = {'涨跌幅': '0.00%', '封单金额': '0.00"亿"', '预测强度': '0.00'}
write_and_style("今日涨停个股", df_main, fmt_main)

# 2. 行业汇总 (按涨停数降序, 强度2位小数)
sector_sum = df.groupby('所属行业').agg(
    涨停数=('股票名称', 'count'),
    最高连板=('连板天数', 'max'),
    平均强度=('预测强度', 'mean')
).reset_index().sort_values('涨停数', ascending=False)
write_and_style("行业汇总", sector_sum, {'平均强度': '0.00'})

# 3. 连板数汇总
board_sum = df.groupby('连板天数').agg(
    个股数量=('股票名称', 'count'),
    个股列表=('股票名称', lambda x: '、'.join(x))
).reset_index().sort_values('连板天数', ascending=False)
write_and_style("连板数汇总", board_sum)

# 4. 三板及以上个股
df_high = df[df['连板天数'] >= 3][['股票代码','股票名称','连板状态','所属行业','预测强度']].sort_values('预测强度', ascending=False)
write_and_style("三板及以上个股", df_high, {'预测强度': '0.00'})

# 5. 预测强度排名
df_rank = df[['股票代码','股票名称','预测强度','强度等级']].sort_values('预测强度', ascending=False)
write_and_style("预测强度排名", df_rank, {'预测强度': '0.00'})

# 保存
file_path = OUTPUT_DIR / f"涨停分析_{today_date}.xlsx"
wb.save(file_path)
print(f"✅ 报告优化完成：{file_path}")